import logging
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Annotated, Any

import dateutil.parser
from fastapi import APIRouter, File, HTTPException, Response
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from pydantic import BaseModel

router = APIRouter(prefix="/api/xlsx", tags=["xlsx"])

_types = {"b": "boolean", "s": "string"}
_log = logging.getLogger(__name__)


def to_type(cell):
    if not cell:
        return "string"
    if cell.is_date:
        return "date"
    if cell.data_type in _types:
        return _types[cell.data_type]
    v = cell.value
    if isinstance(v, (int, int)):
        return "int"
    if isinstance(v, float):
        return "float"
    return "string"


def _convert_value(v):
    if isinstance(v, datetime):
        return v.isoformat()
    return v


class TableColumn(BaseModel):
    name: str
    type: str


CELL_CONTENT = str | int | float | bool | datetime | None


class TableSheet(BaseModel):
    title: str
    columns: list[TableColumn]
    rows: list[dict[str, Any]]


class TableData(BaseModel):
    sheets: list[TableSheet]


@router.post("/to_json/", response_model=TableData)
def xlsx2json(file: Annotated[bytes, File()]):
    if not file:
        raise HTTPException(status_code=403, detail="missing file")

    wb = load_workbook(filename=BytesIO(file), read_only=True, data_only=True)  # type: ignore

    def convert_row(row, cols: list[TableColumn]) -> dict[str, CELL_CONTENT]:
        result = {}

        for r, c in zip(row, cols, strict=False):
            result[c.name] = _convert_value(r.value)
        return result

    def convert_sheet(ws):

        ws_rows = ws.iter_rows()
        ws_cols = next(ws_rows, [])
        ws_first_row = next(ws_rows, [])

        cols = [TableColumn(name=h.value, type=to_type(r)) for h, r in zip(ws_cols, ws_first_row, strict=False)]

        rows = []
        rows.append(convert_row(ws_first_row, cols))
        for row in ws_rows:
            rows.append(convert_row(row, cols))
        return TableSheet(title=ws.title, columns=cols, rows=rows)

    data = TableData(sheets=[convert_sheet(ws) for ws in wb.worksheets])
    return data


@router.post("/to_json_array/", response_model=list[list[Any]])
def xlsx2json_array(file: Annotated[bytes, File()]):
    if not file:
        raise HTTPException(status_code=403, detail="missing file")

    wb = load_workbook(filename=BytesIO(file), read_only=True, data_only=True)

    def convert_row(row):
        return [_convert_value(cell.value) for cell in row]

    if not wb.worksheets:
        return []

    ws = wb.worksheets[0]

    rows = [convert_row(row) for row in ws.iter_rows()]
    return rows


@router.post("/from_json/")
def json2xlsx(data: TableData):
    wb = Workbook(write_only=True)

    bold = Font(bold=True)

    def _escape(v):
        if isinstance(v, str) and v.startswith(("+", "-", "@", "=", "DDE")):
            _log.warning("Escaping possible CSV injection: %s", v)
            return f"'{v}"
        return v

    def to_cell(v):
        v = _escape(v)
        # If the native value cannot be used as Excel value, used the stringified version instead.
        try:
            return WriteOnlyCell(ws, value=v)  # type: ignore
        except ValueError:
            return WriteOnlyCell(ws, value=str(v))  # type: ignore

    def to_header(v):
        c = to_cell(v)
        c.font = bold
        return c

    def to_value(v, coltype):
        if coltype == "date":
            if isinstance(v, int):
                v = datetime.fromtimestamp(v)
            elif isinstance(v, str) and len(v) > 0:
                v = dateutil.parser.parse(v)
        return to_cell(v)

    for sheet in data.sheets:
        ws = wb.create_sheet(title=sheet.title)
        cols = sheet.columns
        ws.append(to_header(col.name) for col in cols)

        for row in sheet.rows:
            ws.append(to_value(row.get(col.name, None), col.type) for col in cols)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        s = tmp.read()
        return Response(
            content=s,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


@router.post("/from_json_array/")
def json_array2xlsx(data: list[list[Any]]):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    for row in data:
        ws.append(row)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        s = tmp.read()
        return Response(
            content=s,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def create():
    """
    entry point of this plugin
    """
    return router
